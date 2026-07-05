# -*- coding: utf-8 -*-
"""特种标签统一产品数据库 —— 轻量原型
查询口子: /            按 品牌/Application/Facestock/Adhesive/耐温/关键词 交叉筛选, 可导出Excel
输入口子: /admin       下载Excel模板 → 填写/修改 → 上传导入(新增+覆盖更新), TDS文件上传绑定单品
运行:     python app/app.py   (默认 http://127.0.0.1:5000)
"""
import io
import os
import re
import sqlite3
from datetime import datetime

from flask import (Flask, g, redirect, render_template, request,
                   send_file, url_for, flash)
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, "data", "labels.db")
UPLOAD_DIR = os.path.join(BASE_DIR, "data", "uploads")
SCHEMA_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "schema.sql")

app = Flask(__name__)
app.secret_key = "etia-pim-prototype"
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB

# ---------------------------------------------------------------- DB helpers

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


SEED_BRANDS = [
    ("3M", "3M", "3M", 1),
    ("BRADY", "Brady", "贝迪", 1),
    ("LINTEC", "LINTEC", "琳得科", 1),
    ("AVERY", "Avery Dennison", "艾利丹尼森", 1),
    ("POLYONICS", "Polyonics", "宝利尼克斯", 1),
    ("YSTECH", "YS Tech Japan", "YS Tech日本", 1),
    ("COMPUTYPE", "Computype", "康普泰", 1),
    ("FLEXCON", "Flexcon", "富林康", 1),
    ("MACTAC", "Mactac", "美泰克", 1),
    ("MARKTECH", "Mark Tech", "马科泰", 1),
    ("ETIA", "ETIA", "ETIA自研", 2),
]
SEED_APPS = [
    ("ELECTRONICS", "电子制造", "Electronics"),
    ("METALLURGY", "冶金钢铁", "Metallurgy"),
    ("AUTOMOTIVE", "汽车制造", "Automotive"),
    ("MEDICAL", "医疗器械", "Medical"),
    ("PHARMACY", "制药", "Pharmacy"),
    ("LAB_BIO", "实验室/生物样本", "Laboratory & Bio"),
    ("ENERGY_CHEM", "能源化工", "Energy & Chemical"),
    ("AEROSPACE", "航空航天", "Aerospace"),
    ("MANUFACTURING", "通用制造", "Manufacturing"),
    ("TRANSPORTATION", "运输物流", "Transportation"),
    ("COMMERCIAL", "商业解决方案", "Commercial Solutions"),
    ("SAFETY", "安全标识", "Safety"),
    ("DATACOM", "数据通信/线缆", "DataCom & Wire ID"),
]
SEED_FACES = [
    ("POLYIMIDE", "PI聚酰亚胺", "Polyimide"),
    ("POLYESTER", "PET聚酯", "Polyester"),
    ("HT_PET", "耐高温PET", "High-Temp Polyester"),
    ("CERAMIC_METAL", "陶瓷/金属基超高温", "Ceramic/Metal Ultra-HT"),
    ("PEN", "PEN", "PEN"),
    ("PPS", "PPS", "PPS"),
    ("ALU_FOIL", "铝箔", "Aluminum Foil"),
    ("PP", "PP聚丙烯", "Polypropylene"),
    ("PE", "PE聚乙烯", "Polyethylene"),
    ("VINYL", "PVC乙烯基", "Vinyl"),
    ("POLYOLEFIN", "聚烯烃", "Polyolefin"),
    ("CLOTH", "醋酸布/玻璃布", "Cloth"),
    ("TYVEK", "Tyvek", "Tyvek"),
    ("PAPER", "纸类", "Paper"),
    ("SYNTHETIC_PAPER", "合成纸", "Synthetic Paper"),
    ("NYLON_CLOTH", "尼龙布", "Nylon Cloth"),
    ("TAMPER_EVIDENT", "防拆易碎", "Tamper Evident"),
]
SEED_ADHS = [
    ("HT_ACRYLIC", "高温丙烯酸胶", "High-Temp Acrylic"),
    ("ACRYLIC", "标准丙烯酸胶", "Acrylic"),
    ("SILICONE", "硅胶系", "Silicone"),
    ("RUBBER", "橡胶系", "Rubber"),
    ("REMOVABLE", "可移除胶", "Removable"),
    ("AGGRESSIVE_PERM", "超强永久胶", "Aggressive Permanent"),
    ("COLD_TEMP", "低温冷贴胶", "Cold Temp"),
    ("MEDICAL_GRADE", "医用低致敏胶", "Medical Grade"),
    ("NONE", "无胶/机械固定", "None/Mechanical"),
]


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    db = sqlite3.connect(DB_PATH)
    with open(SCHEMA_PATH, encoding="utf-8") as f:
        db.executescript(f.read())
    content_schema = os.path.join(os.path.dirname(SCHEMA_PATH), "schema_content.sql")
    if os.path.exists(content_schema):
        with open(content_schema, encoding="utf-8") as f:
            db.executescript(f.read())
    for code, en, cn, btype in SEED_BRANDS:
        db.execute("INSERT OR IGNORE INTO brand(brand_code,brand_name_en,brand_name_cn,brand_type) VALUES(?,?,?,?)",
                   (code, en, cn, btype))
    for code, cn, en in SEED_APPS:
        db.execute("INSERT OR IGNORE INTO application_dict(app_code,app_name_cn,app_name_en) VALUES(?,?,?)", (code, cn, en))
    for code, cn, en in SEED_FACES:
        db.execute("INSERT OR IGNORE INTO facestock_dict(face_code,face_name_cn,face_name_en) VALUES(?,?,?)", (code, cn, en))
    for code, cn, en in SEED_ADHS:
        db.execute("INSERT OR IGNORE INTO adhesive_dict(adh_code,adh_name_cn,adh_name_en) VALUES(?,?,?)", (code, cn, en))
    db.commit()
    db.close()


# ------------------------------------------------------------ dict resolvers

def norm_code(text):
    return re.sub(r"[^A-Z0-9]+", "_", str(text).strip().upper()).strip("_")[:40]


def resolve_dict(db, table, code_col, id_col, value, auto_create=True):
    """按 code / 英文名 / 中文名 匹配字典项; 未匹配时自动新建, 返回id."""
    if value is None or str(value).strip() == "":
        return None
    v = str(value).strip()
    name_en = {"facestock_dict": "face_name_en", "adhesive_dict": "adh_name_en",
               "application_dict": "app_name_en"}[table]
    name_cn = name_en.replace("_en", "_cn")
    row = db.execute(
        f"SELECT {id_col} AS i FROM {table} WHERE UPPER({code_col})=UPPER(?) "
        f"OR UPPER({name_en})=UPPER(?) OR {name_cn}=?", (v, v, v)).fetchone()
    if row:
        return row["i"]
    if not auto_create:
        return None
    cur = db.execute(f"INSERT INTO {table}({code_col},{name_cn},{name_en}) VALUES(?,?,?)",
                     (norm_code(v), v, v))
    return cur.lastrowid


def resolve_brand(db, value):
    if not value or str(value).strip() == "":
        return None
    v = str(value).strip()
    row = db.execute("SELECT brand_id FROM brand WHERE UPPER(brand_code)=UPPER(?) "
                     "OR UPPER(brand_name_en)=UPPER(?) OR brand_name_cn=?", (v, v, v)).fetchone()
    if row:
        return row["brand_id"]
    cur = db.execute("INSERT INTO brand(brand_code,brand_name_en,brand_name_cn) VALUES(?,?,?)",
                     (norm_code(v), v, v))
    return cur.lastrowid


def temp_tier(temp_long_max):
    if temp_long_max is None:
        return None
    t = int(temp_long_max)
    for tier, hi in (("T1", 80), ("T2", 120), ("T3", 180), ("T4", 260), ("T5", 500)):
        if t <= hi:
            return tier
    return "T6"


def to_int(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(round(float(str(v).replace("℃", "").replace("°C", "").strip())))
    except ValueError:
        return None


# ---------------------------------------------------------------- 查询口子

FILTER_SQL = """
SELECT p.*, b.brand_code, b.brand_name_cn, b.brand_type,
       f.face_name_cn, f.face_name_en, a.adh_name_cn, a.adh_name_en,
       (SELECT GROUP_CONCAT(ad.app_name_cn, ' / ') FROM product_application pa
         JOIN application_dict ad ON ad.app_id = pa.app_id
        WHERE pa.product_id = p.product_id) AS apps,
       (SELECT COUNT(*) FROM product_document d WHERE d.product_id = p.product_id) AS doc_count
FROM product p
JOIN brand b ON b.brand_id = p.brand_id
LEFT JOIN facestock_dict f ON f.face_id = p.face_id
LEFT JOIN adhesive_dict a ON a.adh_id = p.adh_id
WHERE p.status = 1 {where}
ORDER BY b.sort_order, b.brand_code, p.model_no
"""


def build_filters(args):
    where, params = "", []
    for key, clause in (("brand", "p.brand_id IN ({ph})"),
                        ("app", "p.product_id IN (SELECT product_id FROM product_application WHERE app_id IN ({ph}))"),
                        ("face", "p.face_id IN ({ph})"),
                        ("adh", "p.adh_id IN ({ph})")):
        vals = [v for v in args.getlist(key) if v]
        if vals:
            where += " AND " + clause.format(ph=",".join("?" * len(vals)))
            params += vals
    tmin = args.get("temp_min", "").strip()
    if tmin:
        where += " AND p.temp_long_max >= ?"
        params.append(int(tmin))
    chem = [v for v in args.getlist("chem") if v]
    if chem:
        where += " AND p.chem_grade IN ({})".format(",".join("?" * len(chem)))
        params += chem
    kw = args.get("kw", "").strip()
    if kw:
        where += (" AND (p.model_no LIKE ? OR p.product_name_cn LIKE ? OR p.product_name_en LIKE ?"
                  " OR p.feature LIKE ? OR p.application_desc LIKE ?)")
        params += [f"%{kw}%"] * 5
    return where, params


def dict_options(db):
    return dict(
        brands=db.execute("SELECT * FROM brand WHERE status=1 ORDER BY sort_order, brand_code").fetchall(),
        apps=db.execute("SELECT * FROM application_dict WHERE status=1 ORDER BY sort_order, app_code").fetchall(),
        faces=db.execute("SELECT * FROM facestock_dict WHERE status=1 ORDER BY sort_order, face_code").fetchall(),
        adhs=db.execute("SELECT * FROM adhesive_dict WHERE status=1 ORDER BY sort_order, adh_code").fetchall(),
    )


@app.route("/")
def search():
    db = get_db()
    where, params = build_filters(request.args)
    rows = db.execute(FILTER_SQL.format(where=where), params).fetchall()
    total = db.execute("SELECT COUNT(*) FROM product WHERE status=1").fetchone()[0]
    return render_template("search.html", rows=rows, total=total, args=request.args, **dict_options(db))


@app.route("/product/<int:pid>")
def product_detail(pid):
    db = get_db()
    row = db.execute(FILTER_SQL.format(where=" AND p.product_id=?"), [pid]).fetchone()
    if not row:
        return "产品不存在", 404
    docs = db.execute("SELECT * FROM product_document WHERE product_id=? ORDER BY doc_type, is_latest DESC", (pid,)).fetchall()
    return render_template("detail.html", p=row, docs=docs)


@app.route("/doc/<int:doc_id>")
def download_doc(doc_id):
    db = get_db()
    doc = db.execute("SELECT * FROM product_document WHERE doc_id=?", (doc_id,)).fetchone()
    if not doc:
        return "文件不存在", 404
    return send_file(os.path.join(UPLOAD_DIR, doc["file_path"]), as_attachment=True,
                     download_name=os.path.basename(doc["file_path"]))


EXPORT_COLS = [  # (Excel表头, 行取值函数)
    ("品牌 brand_code", lambda r: r["brand_code"]),
    ("型号 model_no", lambda r: r["model_no"]),
    ("中文品名", lambda r: r["product_name_cn"]),
    ("英文品名", lambda r: r["product_name_en"]),
    ("面材 facestock", lambda r: r["face_name_en"]),
    ("胶水 adhesive", lambda r: r["adh_name_en"]),
    ("行业应用 applications", lambda r: r["apps"]),
    ("长期耐温min℃", lambda r: r["temp_long_min"]),
    ("长期耐温max℃", lambda r: r["temp_long_max"]),
    ("峰值耐温℃", lambda r: r["temp_peak_max"]),
    ("耐温档位", lambda r: r["temp_tier"]),
    ("耐化学等级", lambda r: r["chem_grade"]),
    ("厚度um", lambda r: r["thickness_um"]),
    ("颜色", lambda r: r["color"]),
    ("打印方式", lambda r: r["print_method"]),
    ("认证", lambda r: r["certification"]),
    ("特性 Feature", lambda r: r["feature"]),
    ("优势 Benefit", lambda r: r["benefit"]),
    ("适用工况 Application", lambda r: r["application_desc"]),
    ("原始分类", lambda r: r["source_category_raw"]),
    ("来源URL", lambda r: r["source_url"]),
]


@app.route("/export")
def export():
    db = get_db()
    where, params = build_filters(request.args)
    rows = db.execute(FILTER_SQL.format(where=where), params).fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "产品清单"
    ws.append([h for h, _ in EXPORT_COLS])
    for r in rows:
        ws.append([fn(r) for _, fn in EXPORT_COLS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name = "标签产品筛选导出_{}.xlsx".format(datetime.now().strftime("%Y%m%d_%H%M"))
    return send_file(buf, as_attachment=True, download_name=name,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------- 输入口子

IMPORT_HEADERS = [
    "brand_code*", "model_no*", "product_name_cn", "product_name_en",
    "facestock", "adhesive", "applications(逗号分隔)",
    "temp_long_min", "temp_long_max", "temp_peak_max", "chem_grade(A/B/C/D)",
    "thickness_um", "color", "print_method", "certification",
    "feature", "benefit", "application_desc",
    "source_category_raw", "source_url", "is_published(1/0)", "tds_url", "sub_application",
]


@app.route("/admin")
def admin():
    db = get_db()
    stats = db.execute(
        "SELECT b.brand_code, COUNT(p.product_id) AS n FROM brand b "
        "LEFT JOIN product p ON p.brand_id=b.brand_id AND p.status=1 "
        "GROUP BY b.brand_id ORDER BY n DESC, b.brand_code").fetchall()
    products = db.execute(
        "SELECT p.product_id, b.brand_code, p.model_no FROM product p "
        "JOIN brand b ON b.brand_id=p.brand_id WHERE p.status=1 "
        "ORDER BY b.brand_code, p.model_no").fetchall()
    return render_template("admin.html", stats=stats, products=products)


@app.route("/admin/template")
def download_template():
    db = get_db()
    wb = Workbook()
    ws = wb.active
    ws.title = "产品导入"
    ws.append(IMPORT_HEADERS)
    ws.append(["3M", "7818", "3M高温聚酯标签", "3M Polyester Label 7818",
               "Polyester", "Acrylic", "Electronics,Manufacturing",
               -40, 149, 260, "B", 50, "White", "Thermal Transfer", "UL969, RoHS",
               "50um哑光PET面材", "耐回流焊过程,条码保持可读", "PCB过程追溯",
               "Labels > Electronics", "https://www.3m.com/...", 1])
    ws2 = wb.create_sheet("字典参考(勿改)")
    ws2.append(["-- 以下值可直接填入对应列; 也可填新值, 系统自动创建字典项 --"])
    for title, q, cols in (
            ("品牌 brand_code", "SELECT brand_code, brand_name_cn FROM brand ORDER BY brand_code", 2),
            ("面材 facestock", "SELECT face_code, face_name_cn, face_name_en FROM facestock_dict", 3),
            ("胶水 adhesive", "SELECT adh_code, adh_name_cn, adh_name_en FROM adhesive_dict", 3),
            ("行业 applications", "SELECT app_code, app_name_cn, app_name_en FROM application_dict", 3)):
        ws2.append([])
        ws2.append([f"== {title} =="])
        for r in get_db().execute(q).fetchall():
            ws2.append(list(r)[:cols])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="产品导入模板.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/import", methods=["POST"])
def import_excel():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("未选择文件")
        return redirect(url_for("admin"))
    db = get_db()
    wb = load_workbook(io.BytesIO(f.read()), data_only=True)
    ws = wb.worksheets[0]
    added = updated = skipped = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row) + [None] * (len(IMPORT_HEADERS) - len(row))
        (brand, model, name_cn, name_en, face, adh, apps_s,
         tmin, tmax, tpeak, chem, thick, color, printm, cert,
         feature, benefit, app_desc, src_cat, src_url, pub, tds_u, sub_app) = vals[:23]
        if not brand or not model:
            skipped += 1
            continue
        try:
            brand_id = resolve_brand(db, brand)
            face_id = resolve_dict(db, "facestock_dict", "face_code", "face_id", face)
            adh_id = resolve_dict(db, "adhesive_dict", "adh_code", "adh_id", adh)
            tmax_i = to_int(tmax)
            chem_v = str(chem).strip().upper()[:1] if chem else None
            data = dict(product_name_cn=name_cn, product_name_en=name_en,
                        face_id=face_id, adh_id=adh_id,
                        temp_long_min=to_int(tmin), temp_long_max=tmax_i,
                        temp_peak_max=to_int(tpeak), temp_tier=temp_tier(tmax_i),
                        chem_grade=chem_v if chem_v in ("A", "B", "C", "D") else None,
                        thickness_um=to_int(thick), color=color, print_method=printm,
                        certification=cert, feature=feature, benefit=benefit,
                        application_desc=app_desc, source_category_raw=src_cat,
                        source_url=src_url, tds_url=tds_u, sub_application=sub_app,
                        is_published=1 if pub in (None, "", 1, "1") else 0)
            exist = db.execute("SELECT product_id FROM product WHERE brand_id=? AND model_no=?",
                               (brand_id, str(model).strip())).fetchone()
            if exist:
                pid = exist["product_id"]
                sets = ", ".join(f"{k}=?" for k in data)
                db.execute(f"UPDATE product SET {sets}, updated_at=datetime('now','localtime') WHERE product_id=?",
                           list(data.values()) + [pid])
                updated += 1
            else:
                cols = "brand_id, model_no, " + ", ".join(data)
                ph = ",".join("?" * (len(data) + 2))
                cur = db.execute(f"INSERT INTO product({cols}) VALUES({ph})",
                                 [brand_id, str(model).strip()] + list(data.values()))
                pid = cur.lastrowid
                added += 1
            db.execute("DELETE FROM product_application WHERE product_id=?", (pid,))
            for one in re.split(r"[,，;；/]+", str(apps_s or "")):
                app_id = resolve_dict(db, "application_dict", "app_code", "app_id", one)
                if app_id:
                    db.execute("INSERT OR IGNORE INTO product_application(product_id,app_id) VALUES(?,?)", (pid, app_id))
        except Exception as e:  # noqa: BLE001 —— 单行失败不中断整体导入
            errors.append(f"第{i}行: {e}")
    db.commit()
    msg = f"导入完成: 新增 {added} 条, 更新 {updated} 条, 跳过空行 {skipped} 条"
    if errors:
        msg += " | 错误{}条: {}".format(len(errors), "; ".join(errors[:5]))
    flash(msg)
    return redirect(url_for("admin"))


@app.route("/admin/upload_doc", methods=["POST"])
def upload_doc():
    pid = request.form.get("product_id")
    f = request.files.get("file")
    if not pid or not f or not f.filename:
        flash("请选择产品和文件")
        return redirect(url_for("admin"))
    db = get_db()
    doc_type = request.form.get("doc_type", "TDS")
    sub = datetime.now().strftime("%Y%m")
    os.makedirs(os.path.join(UPLOAD_DIR, sub), exist_ok=True)
    fname = "{}_{}_{}".format(pid, doc_type, secure_filename(f.filename))
    rel = os.path.join(sub, fname)
    f.save(os.path.join(UPLOAD_DIR, rel))
    if request.form.get("is_latest", "1") == "1":
        db.execute("UPDATE product_document SET is_latest=0 WHERE product_id=? AND doc_type=?", (pid, doc_type))
    db.execute("INSERT INTO product_document(product_id,doc_type,doc_title,file_path,language,version,is_latest,is_public)"
               " VALUES(?,?,?,?,?,?,?,?)",
               (pid, doc_type, request.form.get("doc_title") or f.filename, rel,
                request.form.get("language", "en"), request.form.get("version"),
                1, 1 if request.form.get("is_public") == "1" else 0))
    db.commit()
    flash(f"文件已上传并绑定产品 #{pid}")
    return redirect(url_for("admin"))


# ---------------------------------------------------------------- 单品结构化录入

@app.route("/admin/add_product", methods=["POST"])
def add_product():
    fm = request.form
    brand, model = fm.get("brand", "").strip(), fm.get("model", "").strip()
    if not brand or not model:
        flash("品牌和型号必填")
        return redirect(url_for("admin"))
    db = get_db()
    brand_id = resolve_brand(db, brand)
    tmax_i = to_int(fm.get("tmax"))
    chem_v = (fm.get("chem") or "").strip().upper()[:1]
    data = dict(product_name_cn=fm.get("name_cn"), product_name_en=fm.get("name_en"),
                face_id=resolve_dict(db, "facestock_dict", "face_code", "face_id", fm.get("face")),
                adh_id=resolve_dict(db, "adhesive_dict", "adh_code", "adh_id", fm.get("adh")),
                temp_long_min=to_int(fm.get("tmin")), temp_long_max=tmax_i,
                temp_peak_max=to_int(fm.get("tpeak")), temp_tier=temp_tier(tmax_i),
                chem_grade=chem_v if chem_v in ("A", "B", "C", "D") else None,
                thickness_um=to_int(fm.get("thick")), color=fm.get("color"),
                print_method=fm.get("print_method"), certification=fm.get("cert"),
                feature=fm.get("feature"), benefit=fm.get("benefit"),
                application_desc=fm.get("app_desc"),
                source_category_raw=fm.get("source_raw"), source_url=fm.get("source_url"),
                tds_url=fm.get("tds_url"),
                is_published=1 if fm.get("pub") == "1" else 0)
    exist = db.execute("SELECT product_id FROM product WHERE brand_id=? AND model_no=?",
                       (brand_id, model)).fetchone()
    if exist:
        pid = exist["product_id"]
        sets = ", ".join(f"{k}=?" for k in data)
        db.execute(f"UPDATE product SET {sets}, updated_at=datetime('now','localtime') WHERE product_id=?",
                   list(data.values()) + [pid])
        verb = "更新"
    else:
        cols = "brand_id, model_no, " + ", ".join(data)
        ph = ",".join("?" * (len(data) + 2))
        pid = db.execute(f"INSERT INTO product({cols}) VALUES({ph})",
                         [brand_id, model] + list(data.values())).lastrowid
        verb = "新增"
    db.execute("DELETE FROM product_application WHERE product_id=?", (pid,))
    for one in re.split(r"[,，;；/]+", fm.get("apps", "")):
        app_id = resolve_dict(db, "application_dict", "app_code", "app_id", one)
        if app_id:
            db.execute("INSERT OR IGNORE INTO product_application(product_id,app_id) VALUES(?,?)", (pid, app_id))
    db.commit()
    flash(f"单品{verb}成功: {brand} {model}")
    return redirect(url_for("product_detail", pid=pid))


# ---------------------------------------------------------------- 下载中心 / 定制Catalog

CHROMIUM = "/opt/pw-browsers/chromium"


def html_to_pdf(html_str):
    """打印HTML为PDF字节(用无头Chromium)."""
    import subprocess
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        src = os.path.join(td, "catalog.html")
        out = os.path.join(td, "catalog.pdf")
        with open(src, "w", encoding="utf-8") as f:
            f.write(html_str)
        subprocess.run([CHROMIUM, "--headless", "--no-sandbox", "--disable-gpu",
                        "--no-pdf-header-footer", f"--print-to-pdf={out}",
                        "file://" + src], check=True, capture_output=True, timeout=90)
        with open(out, "rb") as f:
            return f.read()


@app.route("/downloads")
def downloads():
    db = get_db()
    where, params = build_filters(request.args)
    rows = db.execute(FILTER_SQL.format(where=where), params).fetchall()
    return render_template("downloads.html", rows=rows, args=request.args, **dict_options(db))


@app.route("/downloads/create", methods=["POST"])
def downloads_create():
    ids = [int(i) for i in request.form.getlist("pid") if str(i).isdigit()]
    if not ids:
        flash("请先勾选要放进目录的产品")
        return redirect(url_for("downloads"))
    fmt = request.form.get("fmt", "xlsx")
    title = request.form.get("title", "").strip() or "ETIA 定制产品目录"
    db = get_db()
    ph = ",".join("?" * len(ids))
    rows = db.execute(FILTER_SQL.format(where=f" AND p.product_id IN ({ph})"), ids).fetchall()
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    if fmt == "pdf":
        html = render_template("catalog_pdf.html", rows=rows, title=title,
                               date=datetime.now().strftime("%Y-%m-%d"))
        pdf = html_to_pdf(html)
        return send_file(io.BytesIO(pdf), as_attachment=True,
                         download_name=f"{title}_{stamp}.pdf", mimetype="application/pdf")
    wb = Workbook()
    ws = wb.active
    ws.title = "定制目录"
    ws.append([h for h, _ in EXPORT_COLS] + ["TDS链接"])
    for r in rows:
        ws.append([fn(r) for _, fn in EXPORT_COLS] + [r["tds_url"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"{title}_{stamp}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def migrate_db():
    db = sqlite3.connect(DB_PATH)
    for col in ("tds_url", "sub_application"):
        try:
            db.execute(f"ALTER TABLE product ADD COLUMN {col} TEXT")
            db.commit()
        except sqlite3.OperationalError:
            pass
    db.close()


init_db()
migrate_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
